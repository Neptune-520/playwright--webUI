import json
import logging
from datetime import datetime
from io import BytesIO
from typing import Dict, Any, List, Optional

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import (
    ElementLocator, TestScript, TestStep,
    ActionSet, ActionSetStep, ActionSetParameter
)
from core.models import Group

logger = logging.getLogger(__name__)

VALID_SCRIPT_ACTION_TYPES = [choice[0] for choice in TestStep.ACTION_TYPE_CHOICES]
VALID_ACTION_SET_ACTION_TYPES = [choice[0] for choice in ActionSetStep.ACTION_TYPE_CHOICES]
VALID_LOCATOR_TYPES = [choice[0] for choice in ElementLocator.LOCATOR_TYPE_CHOICES]
VALID_ACTION_SET_CATEGORIES = [choice[0] for choice in ActionSet.CATEGORY_CHOICES]
VALID_SCRIPT_STATUSES = [choice[0] for choice in TestScript.STATUS_CHOICES]


class ScriptExportImportService:

    def export_scripts_to_json(self, script_ids: List[int]) -> Dict[str, Any]:
        scripts = TestScript.objects.filter(id__in=script_ids).prefetch_related(
            'steps', 'steps__element', 'steps__action_set_ref', 'group'
        )
        action_sets_collected = {}
        export_data = {
            'export_info': {
                'version': '1.0',
                'exported_at': timezone.now().isoformat(),
                'platform': 'automation_test_platform',
                'script_count': scripts.count(),
                'action_set_count': 0,
            },
            'action_sets': [],
            'scripts': [],
        }

        for script in scripts:
            script_data = self._serialize_script(script)
            export_data['scripts'].append(script_data)

            for step in script.steps.all():
                if step.action_type == 'action_set' and step.action_set_ref:
                    as_code = step.action_set_ref.code
                    if as_code not in action_sets_collected:
                        action_sets_collected[as_code] = self._serialize_action_set(step.action_set_ref)

        export_data['action_sets'] = list(action_sets_collected.values())
        export_data['export_info']['action_set_count'] = len(export_data['action_sets'])

        return export_data

    def export_scripts_to_excel(self, script_ids: List[int]) -> Workbook:
        data = self.export_scripts_to_json(script_ids)
        wb = Workbook()

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        def style_header(ws, headers):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            for col_idx in range(1, len(headers) + 1):
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).border = thin_border

        def write_rows(ws, rows, headers):
            style_header(ws, headers)
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    cell.border = thin_border

        ws_scripts = wb.active
        ws_scripts.title = '脚本概览'
        script_headers = ['脚本代码', '脚本名称', '描述', '状态', '版本', '目标URL', '分组代码', '分组名称']
        script_rows = []
        for s in data['scripts']:
            script_rows.append([
                s['code'], s['name'], s['description'], s['status'],
                s['version'], s['target_url'],
                s.get('group_code', ''), s.get('group_name', ''),
            ])
        write_rows(ws_scripts, script_rows, script_headers)

        ws_steps = wb.create_sheet('脚本步骤')
        step_headers = [
            '脚本代码', '步骤名称', '顺序', '操作类型', '元素代码', '操作值',
            '操作配置(JSON)', '动作集合代码', '动作集合参数(JSON)',
            '描述', '是否启用', '失败时继续', '重试次数', '重试间隔',
        ]
        step_rows = []
        for s in data['scripts']:
            for step in s.get('steps', []):
                step_rows.append([
                    s['code'], step['name'], step['order'], step['action_type'],
                    step.get('element_code', '') or '', step.get('action_value', ''),
                    json.dumps(step.get('action_config', {}), ensure_ascii=False),
                    step.get('action_set_code', '') or '',
                    json.dumps(step.get('action_set_params', {}), ensure_ascii=False),
                    step.get('description', ''), step.get('is_enabled', True),
                    step.get('continue_on_failure', False), step.get('retry_count', 0),
                    step.get('retry_interval', 1000),
                ])
        write_rows(ws_steps, step_rows, step_headers)

        ws_elements = wb.create_sheet('元素定位器')
        elem_headers = [
            '脚本代码', '元素代码', '元素名称', '定位类型', '定位值',
            '页面URL', '描述', '等待超时', '等待状态', '是否启用',
        ]
        elem_rows = []
        for s in data['scripts']:
            for elem in s.get('elements', []):
                elem_rows.append([
                    s['code'], elem['code'], elem['name'], elem['locator_type'],
                    elem['locator_value'], elem.get('page_url', ''),
                    elem.get('description', ''), elem.get('wait_timeout', 10000),
                    elem.get('wait_state', 'visible'), elem.get('is_active', True),
                ])
        write_rows(ws_elements, elem_rows, elem_headers)

        ws_as = wb.create_sheet('动作集合')
        as_headers = [
            '动作集合代码', '名称', '描述', '分类', '分组代码', '分组名称',
            '是否内置', '是否启用',
        ]
        as_rows = []
        for as_data in data.get('action_sets', []):
            as_rows.append([
                as_data['code'], as_data['name'], as_data['description'],
                as_data.get('category', 'general'),
                as_data.get('group_code', ''), as_data.get('group_name', ''),
                as_data.get('is_builtin', False), as_data.get('is_active', True),
            ])
        write_rows(ws_as, as_rows, as_headers)

        ws_as_steps = wb.create_sheet('动作集合步骤')
        as_step_headers = [
            '动作集合代码', '步骤名称', '顺序', '操作类型', '定位类型', '定位值',
            '定位描述', '操作值', '值类型', '参数名称', '等待超时', '失败时继续',
            '重试次数', '重试间隔', '随机选项(JSON)', '选择模式', '随机最小值',
            '随机最大值', '强制点击', '描述', '是否启用',
        ]
        as_step_rows = []
        for as_data in data.get('action_sets', []):
            for step in as_data.get('steps', []):
                as_step_rows.append([
                    as_data['code'], step['name'], step['order'], step['action_type'],
                    step.get('locator_type', ''), step.get('locator_value', ''),
                    step.get('locator_description', ''), step.get('action_value', ''),
                    step.get('action_value_type', 'static'), step.get('parameter_name', ''),
                    step.get('wait_timeout', 10000), step.get('continue_on_failure', False),
                    step.get('retry_count', 0), step.get('retry_interval', 1000),
                    json.dumps(step.get('random_options'), ensure_ascii=False) if step.get('random_options') is not None else '',
                    step.get('select_mode', 'dropdown'),
                    step.get('random_min', ''), step.get('random_max', ''),
                    step.get('force_click', False), step.get('description', ''),
                    step.get('is_enabled', True),
                ])
        write_rows(ws_as_steps, as_step_rows, as_step_headers)

        ws_as_params = wb.create_sheet('动作集合参数')
        as_param_headers = [
            '动作集合代码', '参数名称', '参数代码', '描述', '默认值', '是否必填', '排序',
        ]
        as_param_rows = []
        for as_data in data.get('action_sets', []):
            for param in as_data.get('parameters', []):
                as_param_rows.append([
                    as_data['code'], param['name'], param['code'],
                    param.get('description', ''), param.get('default_value', ''),
                    param.get('is_required', True), param.get('order', 0),
                ])
        write_rows(ws_as_params, as_param_rows, as_param_headers)

        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 4, 50)
                ws.column_dimensions[col[0].column_letter].width = adjusted_width

        return wb

    def import_scripts_from_json(
        self,
        file_data,
        conflict_strategy: str = 'skip',
        user=None,
    ) -> Dict[str, Any]:
        result = {
            'success': 0,
            'failed': 0,
            'skipped': 0,
            'errors': [],
            'action_sets_created': 0,
            'action_sets_skipped': 0,
            'action_sets_overwritten': 0,
            'scripts_created': 0,
            'scripts_skipped': 0,
            'scripts_overwritten': 0,
            'elements_created': 0,
            'elements_reused': 0,
        }

        try:
            if isinstance(file_data, bytes):
                data = json.loads(file_data.decode('utf-8'))
            elif isinstance(file_data, str):
                data = json.loads(file_data)
            else:
                data = file_data
        except (json.JSONDecodeError, UnicodeDecodeError) as e:
            result['errors'].append({'type': 'file_parse', 'message': f'JSON解析失败: {str(e)}'})
            result['failed'] = 1
            return result

        validation = self._validate_import_data(data)
        if not validation['valid']:
            result['errors'].extend(validation['errors'])
            result['failed'] = 1
            return result

        self._import_action_sets(data.get('action_sets', []), conflict_strategy, user, result)
        self._import_scripts(data.get('scripts', []), conflict_strategy, user, result)

        result['success'] = result['scripts_created'] + result['action_sets_created']
        result['skipped'] = result['scripts_skipped'] + result['action_sets_skipped']
        result['failed'] = len(result['errors'])

        return result

    def import_scripts_from_excel(
        self,
        file_data,
        conflict_strategy: str = 'skip',
        user=None,
    ) -> Dict[str, Any]:
        from openpyxl import load_workbook

        try:
            if isinstance(file_data, bytes):
                wb = load_workbook(filename=BytesIO(file_data), read_only=True)
            else:
                wb = load_workbook(filename=file_data, read_only=True)
        except Exception as e:
            return {
                'success': 0, 'failed': 1, 'skipped': 0,
                'errors': [{'type': 'file_parse', 'message': f'Excel文件读取失败: {str(e)}'}],
                'action_sets_created': 0, 'action_sets_skipped': 0, 'action_sets_overwritten': 0,
                'scripts_created': 0, 'scripts_skipped': 0, 'scripts_overwritten': 0,
                'elements_created': 0, 'elements_reused': 0,
            }

        data = self._excel_to_dict(wb)
        wb.close()

        return self.import_scripts_from_json(data, conflict_strategy, user)

    def _serialize_script(self, script: TestScript) -> Dict[str, Any]:
        steps_data = []
        elements_data = []
        elements_seen = set()

        for step in script.steps.all().order_by('order'):
            step_dict = {
                'name': step.name,
                'order': step.order,
                'action_type': step.action_type,
                'element_code': step.element.code if step.element else None,
                'action_value': step.action_value,
                'action_config': step.action_config or {},
                'action_set_code': step.action_set_ref.code if step.action_set_ref else None,
                'action_set_params': step.action_set_params or {},
                'description': step.description,
                'is_enabled': step.is_enabled,
                'continue_on_failure': step.continue_on_failure,
                'retry_count': step.retry_count,
                'retry_interval': step.retry_interval,
            }
            steps_data.append(step_dict)

            if step.element and step.element.code not in elements_seen:
                elements_seen.add(step.element.code)
                elements_data.append({
                    'code': step.element.code,
                    'name': step.element.name,
                    'locator_type': step.element.locator_type,
                    'locator_value': step.element.locator_value,
                    'page_url': step.element.page_url,
                    'description': step.element.description,
                    'wait_timeout': step.element.wait_timeout,
                    'wait_state': step.element.wait_state,
                    'is_active': step.element.is_active,
                })

        return {
            'name': script.name,
            'code': script.code,
            'description': script.description,
            'status': script.status,
            'version': script.version,
            'target_url': script.target_url,
            'group_code': script.group.code if script.group else None,
            'group_name': script.group.name if script.group else None,
            'steps': steps_data,
            'elements': elements_data,
        }

    def _serialize_action_set(self, action_set: ActionSet) -> Dict[str, Any]:
        steps_data = []
        for step in action_set.steps.all().order_by('order'):
            steps_data.append({
                'name': step.name,
                'order': step.order,
                'action_type': step.action_type,
                'locator_type': step.locator_type,
                'locator_value': step.locator_value,
                'locator_description': step.locator_description,
                'action_value': step.action_value,
                'action_value_type': step.action_value_type,
                'parameter_name': step.parameter_name,
                'wait_timeout': step.wait_timeout,
                'continue_on_failure': step.continue_on_failure,
                'retry_count': step.retry_count,
                'retry_interval': step.retry_interval,
                'random_options': step.random_options,
                'select_mode': step.select_mode,
                'random_min': step.random_min,
                'random_max': step.random_max,
                'force_click': step.force_click,
                'description': step.description,
                'is_enabled': step.is_enabled,
            })

        params_data = []
        for param in action_set.parameters.all().order_by('order'):
            params_data.append({
                'name': param.name,
                'code': param.code,
                'description': param.description,
                'default_value': param.default_value,
                'is_required': param.is_required,
                'order': param.order,
            })

        return {
            'name': action_set.name,
            'code': action_set.code,
            'description': action_set.description,
            'category': action_set.category,
            'is_builtin': action_set.is_builtin,
            'is_active': action_set.is_active,
            'group_code': action_set.group.code if action_set.group else None,
            'group_name': action_set.group.name if action_set.group else None,
            'steps': steps_data,
            'parameters': params_data,
        }

    def _validate_import_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        errors = []

        if not isinstance(data, dict):
            errors.append({'type': 'format', 'message': '导入数据格式错误：期望JSON对象'})
            return {'valid': False, 'errors': errors}

        if 'scripts' not in data or not isinstance(data['scripts'], list):
            errors.append({'type': 'format', 'message': '缺少scripts数组或格式不正确'})

        for i, script in enumerate(data.get('scripts', [])):
            script_prefix = f'脚本[{i}]'
            for field in ['name', 'code', 'target_url']:
                if not script.get(field):
                    errors.append({
                        'type': 'validation',
                        'message': f'{script_prefix}: 缺少必填字段 "{field}"',
                    })
            if script.get('status') and script['status'] not in VALID_SCRIPT_STATUSES:
                errors.append({
                    'type': 'validation',
                    'message': f'{script_prefix}: 无效的状态值 "{script["status"]}"',
                })
            for j, step in enumerate(script.get('steps', [])):
                step_prefix = f'{script_prefix}.步骤[{j}]'
                if not step.get('name'):
                    errors.append({
                        'type': 'validation',
                        'message': f'{step_prefix}: 缺少步骤名称',
                    })
                if step.get('action_type') and step['action_type'] not in VALID_SCRIPT_ACTION_TYPES:
                    errors.append({
                        'type': 'validation',
                        'message': f'{step_prefix}: 无效的操作类型 "{step["action_type"]}"',
                    })

        for i, as_data in enumerate(data.get('action_sets', [])):
            as_prefix = f'动作集合[{i}]'
            for field in ['name', 'code']:
                if not as_data.get(field):
                    errors.append({
                        'type': 'validation',
                        'message': f'{as_prefix}: 缺少必填字段 "{field}"',
                    })
            if as_data.get('category') and as_data['category'] not in VALID_ACTION_SET_CATEGORIES:
                errors.append({
                    'type': 'validation',
                    'message': f'{as_prefix}: 无效的分类值 "{as_data["category"]}"',
                })
            param_codes = set()
            for j, param in enumerate(as_data.get('parameters', [])):
                if param.get('code'):
                    if param['code'] in param_codes:
                        errors.append({
                            'type': 'validation',
                            'message': f'{as_prefix}.参数[{j}]: 参数代码 "{param["code"]}" 重复',
                        })
                    param_codes.add(param['code'])
            for j, step in enumerate(as_data.get('steps', [])):
                step_prefix = f'{as_prefix}.步骤[{j}]'
                if step.get('action_type') and step['action_type'] not in VALID_ACTION_SET_ACTION_TYPES:
                    errors.append({
                        'type': 'validation',
                        'message': f'{step_prefix}: 无效的操作类型 "{step["action_type"]}"',
                    })

        return {'valid': len(errors) == 0, 'errors': errors}

    def _import_action_sets(self, action_sets_data, conflict_strategy, user, result):
        for as_data in action_sets_data:
            code = as_data.get('code')
            if not code:
                result['errors'].append({
                    'type': 'action_set',
                    'message': f'动作集合缺少code字段，已跳过',
                })
                continue

            if as_data.get('is_builtin', False):
                if ActionSet.objects.filter(code=code).exists():
                    result['action_sets_skipped'] += 1
                    continue

            existing = ActionSet.objects.filter(code=code).first()

            if existing:
                if existing.is_builtin:
                    result['action_sets_skipped'] += 1
                    continue
                if conflict_strategy == 'skip':
                    result['action_sets_skipped'] += 1
                    continue
                elif conflict_strategy == 'overwrite':
                    existing.steps.all().delete()
                    existing.parameters.all().delete()
                    existing.name = as_data.get('name', existing.name)
                    existing.description = as_data.get('description', '')
                    existing.category = as_data.get('category', 'general')
                    existing.is_active = as_data.get('is_active', True)
                    group = self._find_group(as_data.get('group_code'), 'action_set')
                    if group:
                        existing.group = group
                    existing.save()
                    self._create_action_set_steps(existing, as_data.get('steps', []))
                    self._create_action_set_parameters(existing, as_data.get('parameters', []))
                    result['action_sets_overwritten'] += 1
                    result['action_sets_created'] += 1
                    continue
                elif conflict_strategy == 'rename':
                    code = self._generate_unique_code(ActionSet, code)

            group = self._find_group(as_data.get('group_code'), 'action_set')
            action_set = ActionSet.objects.create(
                name=as_data.get('name', ''),
                code=code,
                description=as_data.get('description', ''),
                category=as_data.get('category', 'general'),
                is_builtin=False,
                is_active=as_data.get('is_active', True),
                group=group,
                created_by=user,
            )
            self._create_action_set_steps(action_set, as_data.get('steps', []))
            self._create_action_set_parameters(action_set, as_data.get('parameters', []))
            result['action_sets_created'] += 1

    def _import_scripts(self, scripts_data, conflict_strategy, user, result):
        for script_data in scripts_data:
            code = script_data.get('code')
            name = script_data.get('name')
            target_url = script_data.get('target_url')

            if not code or not name or not target_url:
                result['errors'].append({
                    'type': 'script',
                    'message': f'脚本缺少必填字段(name/code/target_url)，已跳过',
                })
                continue

            existing = TestScript.objects.filter(code=code).first()

            if existing:
                if conflict_strategy == 'skip':
                    result['scripts_skipped'] += 1
                    continue
                elif conflict_strategy == 'overwrite':
                    existing.steps.all().delete()
                    existing.name = name
                    existing.description = script_data.get('description', '')
                    existing.target_url = target_url
                    existing.status = script_data.get('status', 'draft')
                    group = self._find_group(script_data.get('group_code'), 'script')
                    if group:
                        existing.group = group
                    existing.save()
                    self._create_script_steps(existing, script_data.get('steps', []), result)
                    result['scripts_overwritten'] += 1
                    result['scripts_created'] += 1
                    continue
                elif conflict_strategy == 'rename':
                    code = self._generate_unique_code(TestScript, code)

            group = self._find_group(script_data.get('group_code'), 'script')
            script = TestScript.objects.create(
                name=name,
                code=code,
                description=script_data.get('description', ''),
                status=script_data.get('status', 'draft'),
                version=script_data.get('version', 1),
                target_url=target_url,
                script_data={},
                group=group,
                created_by=user,
            )
            self._create_script_steps(script, script_data.get('steps', []), result)
            result['scripts_created'] += 1

    def _create_action_set_steps(self, action_set, steps_data):
        for step_data in steps_data:
            ActionSetStep.objects.create(
                action_set=action_set,
                name=step_data.get('name', ''),
                order=step_data.get('order', 0),
                action_type=step_data.get('action_type', 'click'),
                locator_type=step_data.get('locator_type', 'css'),
                locator_value=step_data.get('locator_value', ''),
                locator_description=step_data.get('locator_description', ''),
                action_value=step_data.get('action_value', ''),
                action_value_type=step_data.get('action_value_type', 'static'),
                parameter_name=step_data.get('parameter_name', ''),
                wait_timeout=step_data.get('wait_timeout', 10000),
                continue_on_failure=step_data.get('continue_on_failure', False),
                retry_count=step_data.get('retry_count', 0),
                retry_interval=step_data.get('retry_interval', 1000),
                random_options=step_data.get('random_options'),
                select_mode=step_data.get('select_mode', 'dropdown'),
                random_min=step_data.get('random_min'),
                random_max=step_data.get('random_max'),
                force_click=step_data.get('force_click', False),
                description=step_data.get('description', ''),
                is_enabled=step_data.get('is_enabled', True),
            )

    def _create_action_set_parameters(self, action_set, params_data):
        for param_data in params_data:
            ActionSetParameter.objects.create(
                action_set=action_set,
                name=param_data.get('name', ''),
                code=param_data.get('code', ''),
                description=param_data.get('description', ''),
                default_value=param_data.get('default_value', ''),
                is_required=param_data.get('is_required', True),
                order=param_data.get('order', 0),
            )

    def _create_script_steps(self, script, steps_data, result):
        for step_data in steps_data:
            element = None
            element_code = step_data.get('element_code')
            if element_code:
                element = self._resolve_element(element_code, step_data, script, result)

            action_set_ref = None
            action_set_code = step_data.get('action_set_code')
            if action_set_code:
                action_set_ref = ActionSet.objects.filter(code=action_set_code).first()
                if not action_set_ref:
                    result['errors'].append({
                        'type': 'warning',
                        'message': f'脚本"{script.name}"步骤"{step_data.get("name")}": '
                                   f'引用的动作集合"{action_set_code}"不存在，action_set_ref将为空',
                    })

            TestStep.objects.create(
                script=script,
                name=step_data.get('name', ''),
                order=step_data.get('order', 0),
                action_type=step_data.get('action_type', 'click'),
                element=element,
                action_value=step_data.get('action_value', ''),
                action_config=step_data.get('action_config', {}),
                action_set_ref=action_set_ref,
                action_set_params=step_data.get('action_set_params', {}),
                description=step_data.get('description', ''),
                is_enabled=step_data.get('is_enabled', True),
                continue_on_failure=step_data.get('continue_on_failure', False),
                retry_count=step_data.get('retry_count', 0),
                retry_interval=step_data.get('retry_interval', 1000),
            )

    def _resolve_element(self, element_code, step_data, script, result):
        existing = ElementLocator.objects.filter(code=element_code).first()
        if existing:
            result['elements_reused'] += 1
            return existing

        elements_in_script = None
        if hasattr(script, '_import_elements'):
            elements_in_script = script._import_elements.get(element_code)

        if elements_in_script:
            elem = ElementLocator.objects.create(
                code=element_code,
                name=elements_in_script.get('name', element_code),
                locator_type=elements_in_script.get('locator_type', 'css'),
                locator_value=elements_in_script.get('locator_value', ''),
                page_url=elements_in_script.get('page_url', ''),
                description=elements_in_script.get('description', ''),
                wait_timeout=elements_in_script.get('wait_timeout', 10000),
                wait_state=elements_in_script.get('wait_state', 'visible'),
                is_active=elements_in_script.get('is_active', True),
            )
            result['elements_created'] += 1
            return elem

        return None

    def _find_group(self, group_code, group_type):
        if not group_code:
            return None
        return Group.objects.filter(code=group_code, type=group_type).first()

    def _generate_unique_code(self, model, base_code):
        code = base_code
        counter = 1
        while model.objects.filter(code=code).exists():
            code = f'{base_code}_import_{counter}'
            counter += 1
        return code

    def _excel_to_dict(self, wb) -> Dict[str, Any]:
        data = {'export_info': {}, 'action_sets': [], 'scripts': []}

        ws_scripts = wb.get('脚本概览')
        if ws_scripts:
            scripts_map = {}
            for row in ws_scripts.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                code = str(row[0])
                scripts_map[code] = {
                    'name': str(row[1] or ''),
                    'code': code,
                    'description': str(row[2] or ''),
                    'status': str(row[3] or 'draft'),
                    'version': int(row[4] or 1),
                    'target_url': str(row[5] or ''),
                    'group_code': str(row[6] or '') if row[6] else None,
                    'group_name': str(row[7] or '') if row[7] else None,
                    'steps': [],
                    'elements': [],
                }
            data['scripts'] = list(scripts_map.values())

        ws_steps = wb.get('脚本步骤')
        if ws_steps:
            scripts_by_code = {s['code']: s for s in data['scripts']}
            for row in ws_steps.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                script_code = str(row[0])
                if script_code not in scripts_by_code:
                    continue
                action_config = {}
                if row[6]:
                    try:
                        action_config = json.loads(str(row[6]))
                    except (json.JSONDecodeError, TypeError):
                        pass
                action_set_params = {}
                if row[8]:
                    try:
                        action_set_params = json.loads(str(row[8]))
                    except (json.JSONDecodeError, TypeError):
                        pass
                scripts_by_code[script_code]['steps'].append({
                    'name': str(row[1] or ''),
                    'order': int(row[2] or 0),
                    'action_type': str(row[3] or 'click'),
                    'element_code': str(row[4] or '') if row[4] else None,
                    'action_value': str(row[5] or ''),
                    'action_config': action_config,
                    'action_set_code': str(row[7] or '') if row[7] else None,
                    'action_set_params': action_set_params,
                    'description': str(row[9] or ''),
                    'is_enabled': bool(row[10]) if row[10] is not None else True,
                    'continue_on_failure': bool(row[11]) if row[11] is not None else False,
                    'retry_count': int(row[12] or 0),
                    'retry_interval': int(row[13] or 1000),
                })

        ws_elements = wb.get('元素定位器')
        if ws_elements:
            scripts_by_code = {s['code']: s for s in data['scripts']}
            for row in ws_elements.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                script_code = str(row[0])
                if script_code not in scripts_by_code:
                    continue
                elem_code = str(row[1] or '')
                if not elem_code:
                    continue
                scripts_by_code[script_code]['elements'].append({
                    'code': elem_code,
                    'name': str(row[2] or ''),
                    'locator_type': str(row[3] or 'css'),
                    'locator_value': str(row[4] or ''),
                    'page_url': str(row[5] or ''),
                    'description': str(row[6] or ''),
                    'wait_timeout': int(row[7] or 10000),
                    'wait_state': str(row[8] or 'visible'),
                    'is_active': bool(row[9]) if row[9] is not None else True,
                })

        ws_as = wb.get('动作集合')
        as_map = {}
        if ws_as:
            for row in ws_as.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                code = str(row[0])
                as_map[code] = {
                    'name': str(row[1] or ''),
                    'code': code,
                    'description': str(row[2] or ''),
                    'category': str(row[3] or 'general'),
                    'group_code': str(row[4] or '') if row[4] else None,
                    'group_name': str(row[5] or '') if row[5] else None,
                    'is_builtin': bool(row[6]) if row[6] is not None else False,
                    'is_active': bool(row[7]) if row[7] is not None else True,
                    'steps': [],
                    'parameters': [],
                }
            data['action_sets'] = list(as_map.values())

        ws_as_steps = wb.get('动作集合步骤')
        if ws_as_steps:
            for row in ws_as_steps.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                as_code = str(row[0])
                if as_code not in as_map:
                    continue
                random_options = None
                if row[14]:
                    try:
                        random_options = json.loads(str(row[14]))
                    except (json.JSONDecodeError, TypeError):
                        pass
                as_map[as_code]['steps'].append({
                    'name': str(row[1] or ''),
                    'order': int(row[2] or 0),
                    'action_type': str(row[3] or 'click'),
                    'locator_type': str(row[4] or 'css'),
                    'locator_value': str(row[5] or ''),
                    'locator_description': str(row[6] or ''),
                    'action_value': str(row[7] or ''),
                    'action_value_type': str(row[8] or 'static'),
                    'parameter_name': str(row[9] or ''),
                    'wait_timeout': int(row[10] or 10000),
                    'continue_on_failure': bool(row[11]) if row[11] is not None else False,
                    'retry_count': int(row[12] or 0),
                    'retry_interval': int(row[13] or 1000),
                    'random_options': random_options,
                    'select_mode': str(row[15] or 'dropdown'),
                    'random_min': int(row[16]) if row[16] is not None else None,
                    'random_max': int(row[17]) if row[17] is not None else None,
                    'force_click': bool(row[18]) if row[18] is not None else False,
                    'description': str(row[19] or ''),
                    'is_enabled': bool(row[20]) if row[20] is not None else True,
                })

        ws_as_params = wb.get('动作集合参数')
        if ws_as_params:
            for row in ws_as_params.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                as_code = str(row[0])
                if as_code not in as_map:
                    continue
                as_map[as_code]['parameters'].append({
                    'name': str(row[1] or ''),
                    'code': str(row[2] or ''),
                    'description': str(row[3] or ''),
                    'default_value': str(row[4] or ''),
                    'is_required': bool(row[5]) if row[5] is not None else True,
                    'order': int(row[6] or 0),
                })

        for script in data['scripts']:
            script['_import_elements'] = {e['code']: e for e in script.get('elements', [])}

        return data
